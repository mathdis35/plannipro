import os, re, shutil, datetime, uuid, json, copy, calendar
from pathlib import Path
from collections import defaultdict
from flask import Flask, render_template, request, jsonify, send_file, after_this_request

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    import xlrd
except ImportError:
    pass

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024
UPLOAD_FOLDER = '/tmp/plannipro'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

MOIS_ABBR = {
    'août':8,'aout':8,'sept':9,'septembre':9,'oct':10,'octobre':10,
    'nov':11,'novembre':11,'déc':12,'dec':12,'décembre':12,
    'janv':1,'jan':1,'janvier':1,'févr':2,'fev':2,'février':2,
    'mars':3,'avr':4,'avril':4,'mai':5,'juin':6,'juil':7,'juillet':7
}
MOIS_FR = ['','Janvier','Février','Mars','Avril','Mai','Juin',
           'Juillet','Août','Septembre','Octobre','Novembre','Décembre']
MOIS_FR_UPPER = {
    1:'JANVIER',2:'FÉVRIER',3:'MARS',4:'AVRIL',5:'MAI',6:'JUIN',
    7:'JUILLET',8:'AOÛT',9:'SEPTEMBRE',10:'OCTOBRE',11:'NOVEMBRE',12:'DÉCEMBRE'
}

FERIES = {
    datetime.date(2026,11,1), datetime.date(2026,11,11), datetime.date(2026,12,25),
    datetime.date(2027,1,1),  datetime.date(2027,4,5),   datetime.date(2027,5,1),
    datetime.date(2027,5,8),  datetime.date(2027,5,13),  datetime.date(2027,5,24),
    datetime.date(2027,7,14), datetime.date(2027,8,15),
}

DEFAULT_COLORS = ['FFEE7E32','FFFFCC99','FFFFD243','FFDDFFDD',
                  'FFB8D4F0','FFFFC0CB','FFD4B8F0','FFB8F0D4']


# ─── Paires de colonnes fusionnées dans le template ──────────────────────────
# (découvertes par analyse du template vierge — ne pas modifier)
ALL_MERGE_PAIRS = [
    (4,5),(7,8),(10,11),(13,14),(16,17),(19,20),
    (28,29),(31,32),(34,35),(37,38),(40,41),(43,44),(46,47),(49,50),
    (56,57),(59,60),(62,63),(65,66),(68,69),(71,72),
    (78,79),(81,82),(84,85),(87,88),(90,91)
]
JOURS_COLS   = [2, 26, 54, 76]
JOURS_FR_LIST = ['Lundi','Mardi','Mercredi','Jeudi','Vendredi']

# ─── Helpers ──────────────────────────────────────────────────────────────────
def colors_match(rgb):
    if not rgb: return False
    return abs(rgb[0]-166)<20 and abs(rgb[1]-202)<20 and abs(rgb[2]-240)<20

def cell_to_str(v):
    if v is None: return ''
    return str(v).strip()

def is_available(v):
    if v is None: return False
    return str(v).strip().upper() in ['X','4','✓','OUI','O']

def find_month_num(text):
    t = text.strip().lower()
    for m, mn in MOIS_ABBR.items():
        if m in t:
            yr = re.search(r'(20\d\d)', t)
            return mn, int(yr.group(1)) if yr else None
    return None, None

# ─── Parser planning classe XLS ───────────────────────────────────────────────
def parse_planning_xls(filepath):
    try:
        wb = xlrd.open_workbook(filepath, formatting_info=True)
    except Exception as e:
        return {'nom': Path(filepath).stem, 'jours': []}
    ws = wb.sheet_by_index(0)
    nom = Path(filepath).stem
    for c in range(ws.ncols):
        v = ws.cell_value(0, c)
        if isinstance(v, str) and len(v.strip()) > 3:
            nom = v.strip(); break

    jours = []
    blocks = []
    for c in range(ws.ncols):
        v = ws.cell_value(4, c)
        if isinstance(v, str):
            mn, yr = find_month_num(v)
            if mn and yr:
                blocks.append({'cj': c+1, 'cd': c+2, 'y': yr, 'm': mn})

    for b in blocks:
        for r in range(5, ws.nrows):
            if b['cj'] >= ws.ncols or b['cd'] >= ws.ncols: continue
            try:
                xf = wb.xf_list[ws.cell_xf_index(r, b['cj'])]
                rgb = wb.colour_map.get(xf.background.pattern_colour_index)
                if not colors_match(rgb): continue
            except: continue
            dv = ws.cell_value(r, b['cd'])
            dn = int(dv) if isinstance(dv, float) and dv > 0 else None
            if isinstance(dv, str) and dv.strip().isdigit(): dn = int(dv.strip())
            if not dn or not (1 <= dn <= 31): continue
            try:
                d = datetime.date(b['y'], b['m'], dn)
                if d.weekday() < 5: jours.append(d.strftime('%Y-%m-%d'))
            except: pass
    return {'nom': nom, 'jours': sorted(set(jours))}

# ─── Parser planning classe XLSX ──────────────────────────────────────────────
def parse_planning_xlsx(filepath):
    try:
        wb = load_workbook(filepath, data_only=True)
    except: return {'nom': Path(filepath).stem, 'jours': []}
    ws = wb.active
    nom = Path(filepath).stem
    for c in range(1, ws.max_column+1):
        v = ws.cell(row=1, column=c).value
        if v and isinstance(v, str) and len(v.strip()) > 3:
            nom = v.strip(); break

    jours = []
    blocks = []
    for c in range(1, ws.max_column+1):
        v = ws.cell(row=5, column=c).value
        if isinstance(v, str):
            mn, yr = find_month_num(v)
            if mn and yr:
                blocks.append({'cj': c+1, 'cd': c+2, 'y': yr, 'm': mn})

    for b in blocks:
        for r in range(6, ws.max_row+1):
            cell = ws.cell(row=r, column=b['cj'])
            bg = cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor and cell.fill.fgColor.type == 'rgb' else None
            if bg not in ('FFA6CAF0', 'A6CAF0'): continue
            dv = ws.cell(row=r, column=b['cd']).value
            dn = None
            if isinstance(dv, datetime.datetime): dn = dv.day
            elif isinstance(dv, (int, float)): dn = int(dv)
            elif isinstance(dv, str) and dv.strip().isdigit(): dn = int(dv.strip())
            if not dn or not (1 <= dn <= 31): continue
            try:
                d = datetime.date(b['y'], b['m'], dn)
                if d.weekday() < 5: jours.append(d.strftime('%Y-%m-%d'))
            except: pass
    return {'nom': nom, 'jours': sorted(set(jours))}

def parse_planning_classe(fp):
    return parse_planning_xls(fp) if Path(fp).suffix.lower() == '.xls' else parse_planning_xlsx(fp)

# ─── Parser disponibilités ────────────────────────────────────────────────────
def parse_disponibilite(filepath):
    nom = Path(filepath).stem
    dispo = {}
    try:
        wb = load_workbook(filepath, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except:
        return {'nom': nom, 'dispo': {}}

    if len(rows) < 2: return {'nom': nom, 'dispo': {}}

    header = rows[0]
    month_cols = []
    for ci, val in enumerate(header):
        if isinstance(val, str):
            mn, yr = find_month_num(val)
            if mn and yr:
                month_cols.append({'ci': ci, 'm': mn, 'y': yr})

    ign = {'sam', 'dim', 'férié', 'ferie', 'férie', 'fériés', 'nan', 'none', ''}

    for mc in month_cols:
        ci, mn, yr = mc['ci'], mc['m'], mc['y']
        col_matin = ci
        col_pm    = ci + 1
        for ri in range(2, len(rows)):
            row = rows[ri]
            day_abbr = cell_to_str(row[0]).lower() if len(row) > 0 else ''
            if day_abbr in ign: continue
            day_num_raw = row[1] if len(row) > 1 else None
            dn = None
            if isinstance(day_num_raw, (int, float)) and day_num_raw == day_num_raw:
                dn = int(day_num_raw)
            elif isinstance(day_num_raw, str) and day_num_raw.strip().isdigit():
                dn = int(day_num_raw.strip())
            if not dn or not (1 <= dn <= 31): continue
            try:
                d = datetime.date(yr, mn, dn)
                if d.weekday() >= 5: continue
                ds = d.strftime('%Y-%m-%d')
                mv = row[col_matin] if col_matin < len(row) else None
                pv = row[col_pm]    if col_pm    < len(row) else None
                if cell_to_str(mv).lower() in ign: continue
                dispo[ds] = {'matin': is_available(mv), 'pm': is_available(pv)}
            except: pass

    return {'nom': nom, 'dispo': dispo}

# ─── Parser tableau formateurs ────────────────────────────────────────────────
def parse_tableau_formateurs(filepath):
    assignments = defaultdict(list)

    def process(rows):
        cur = None
        for row in rows:
            v0 = cell_to_str(row[0]) if row and row[0] is not None else ''
            if re.match(r'(BTS|BAC|CGC|NDRC|GPME|Master|RDC|RH|EC)\s', v0, re.I):
                cur = v0
            elif cur and v0 and len(row) > 1 and row[1]:
                mat = cell_to_str(row[1])
                heures_raw = row[-1] if row else 0
                try: h = float(str(heures_raw).split('+')[0].strip())
                except: h = 0
                if h > 0 and mat and mat.lower() not in ['nan', 'total', '']:
                    assignments[cur].append({'formateur': v0, 'matiere': mat, 'heures': h, 'heures_faites': 0})

    ext = Path(filepath).suffix.lower()
    if ext == '.xls':
        try:
            wb = xlrd.open_workbook(filepath)
            si = next((i for i, n in enumerate(wb.sheet_names()) if 'mois' in n.lower()), len(wb.sheet_names())-1)
            ws = wb.sheet_by_index(si)
            process([[ws.cell_value(r, c) for c in range(ws.ncols)] for r in range(ws.nrows)])
        except: pass
    else:
        try:
            wb = load_workbook(filepath, data_only=True)
            sn = next((s for s in wb.sheetnames if 'mois' in s.lower()), wb.sheetnames[-1])
            process(list(wb[sn].iter_rows(values_only=True)))
        except: pass
    return dict(assignments)

# ─── Moteur assignation ───────────────────────────────────────────────────────
def assigner(planning_classes, dispos_formateurs, affectations):
    dispo_idx = {d['nom']: d['dispo'] for d in dispos_formateurs}
    result = defaultdict(dict)
    stats = {'assigned': 0, 'warn': 0}

    for ci in planning_classes:
        cn = ci['nom']; jours = ci['jours']
        aff = next((v for k, v in affectations.items()
                    if k.strip().lower() in cn.lower() or cn.lower() in k.strip().lower()), None)
        if not aff:
            for j in jours: result[j][cn] = {'formateur': '?', 'matiere': '?', 'slot': 'matin'}
            continue
        si = 0
        for j in jours:
            slot = ['matin', 'pm'][si % 2]; si += 1; assigned = None
            for e in sorted(aff, key=lambda x: x['heures_faites']):
                pd_ = dispo_idx.get(e['formateur'], {}); jd = pd_.get(j, {})
                if e['heures'] - e['heures_faites'] <= 0: continue
                if not pd_ or jd.get(slot, False): assigned = e; break
            if assigned:
                assigned['heures_faites'] += 4
                result[j][cn] = {'formateur': assigned['formateur'], 'matiere': assigned['matiere'], 'slot': slot}
                stats['assigned'] += 1
            else:
                result[j][cn] = {'formateur': '⚠️', 'matiere': '', 'slot': slot}
                stats['warn'] += 1

    heures = defaultdict(lambda: defaultdict(int))
    for dv in result.values():
        for cl, inf in dv.items():
            if inf['formateur'] not in ['?', '⚠️']:
                heures[inf['formateur']][inf['matiere']] += 4
    return dict(result), stats, dict(heures)

# ─── Écriture planning assigné ────────────────────────────────────────────────
def ecrire_planning(template_path, assignment, mois_cibles, output_path):
    shutil.copy(template_path, output_path)
    wb = load_workbook(output_path)
    kw = ['BTS','BAC','EC ','CGC','NDRC','GPME','RDC','RH','Master']
    jf = {'lundi':0,'mardi':1,'mercredi':2,'jeudi':3,'vendredi':4,'lun':0,'mar':1,'mer':2,'jeu':3,'ven':4}

    for sn in wb.sheetnames:
        ws = wb[sn]; titre = ''
        for ri in range(1, 4):
            for ci in range(1, ws.max_column+1):
                v = ws.cell(row=ri, column=ci).value
                if v and isinstance(v, str) and len(v.strip()) > 3:
                    titre = v.upper(); break
        mn, yr = None, None
        for m, mnum in MOIS_ABBR.items():
            if m.upper() in titre:
                y = re.search(r'(20\d\d)', titre)
                if y: yr = int(y.group(1)); mn = mnum; break
        if not mn or f'{mn}/{yr}' not in mois_cibles: continue

        cc = {}
        for ri in range(1, 6):
            for ci in range(1, ws.max_column+1):
                v = ws.cell(row=ri, column=ci).value
                if isinstance(v, str) and any(k in v for k in kw): cc[v.strip()] = ci

        dr = {}
        for ri in range(1, ws.max_row+1):
            for ci in range(1, 4):
                v = ws.cell(row=ri, column=ci).value
                if isinstance(v, str) and v.strip().lower() in jf:
                    for lri in [ri, ri+1]:
                        for lci in range(1, 4):
                            dv = ws.cell(row=lri, column=lci).value
                            if isinstance(dv, (int, float)) and 1 <= int(dv) <= 31:
                                try:
                                    d = datetime.date(yr, mn, int(dv))
                                    ds = d.strftime('%Y-%m-%d')
                                    if ds not in dr: dr[ds] = ri
                                except: pass

        for ds, ri in dr.items():
            if ds not in assignment: continue
            for cn, ci in cc.items():
                for k, v in assignment[ds].items():
                    if k.strip().lower() in cn.lower() or cn.lower() in k.strip().lower():
                        cell = ws.cell(row=ri, column=ci)
                        cell.value = '⚠️' if v['formateur'] == '⚠️' else f"{v['formateur']}\n{v['matiere']}"
                        old = cell.font
                        cell.font = Font(name=old.name or 'Calibri', size=old.size or 9, bold=True, color=old.color)
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        break
    wb.save(output_path)

# ─── Génération template colorié (ancien) ────────────────────────────────────
def detect_structure(ws):
    kw = ['BTS','BAC','EC ','CGC','NDRC','GPME','RDC','RH','Master']
    cc = {}; colors = {}
    for ri in range(1, 6):
        for ci in range(1, ws.max_column+1):
            cell = ws.cell(row=ri, column=ci); v = cell.value
            if isinstance(v, str) and any(k in v for k in kw):
                nom = v.strip(); cc[nom] = ci
                if cell.fill and cell.fill.fgColor and cell.fill.fgColor.type == 'rgb':
                    c = cell.fill.fgColor.rgb
                    if c not in ('00000000', 'FFFFFFFF'): colors[nom] = c
    for nom, ci in cc.items():
        if nom in colors: continue
        for ri in range(5, min(25, ws.max_row+1)):
            cell = ws.cell(row=ri, column=ci)
            if cell.fill and cell.fill.fgColor and cell.fill.fgColor.type == 'rgb':
                c = cell.fill.fgColor.rgb
                if c not in ('00000000', 'FFFFFFFF', None):
                    colors[nom] = c; break
    jf_set = {'lundi','mardi','mercredi','jeudi','vendredi','lun','mar','mer','jeu','ven'}
    fdr = 6; dlc = 2
    for ri in range(3, 20):
        for ci in range(1, 5):
            v = ws.cell(row=ri, column=ci).value
            if isinstance(v, str) and v.strip().lower() in jf_set:
                fdr = ri; dlc = ci; break
        else: continue
        break
    return {'class_cols': cc, 'class_colors': colors, 'first_data_row': fdr, 'day_label_col': dlc}

def generer_template_colorie(template_path, planning_classes, annee_debut, output_path):
    wb_tpl = load_workbook(template_path)
    struct = detect_structure(wb_tpl.active)
    cc = struct['class_cols']; colors = struct['class_colors']
    fdr = struct['first_data_row']; dlc = struct['day_label_col']

    ci = 0
    for nom in cc:
        if nom not in colors:
            colors[nom] = DEFAULT_COLORS[ci % len(DEFAULT_COLORS)]; ci += 1

    mois_scolaire = [(9,annee_debut),(10,annee_debut),(11,annee_debut),(12,annee_debut),
                     (1,annee_debut+1),(2,annee_debut+1),(3,annee_debut+1),(4,annee_debut+1),
                     (5,annee_debut+1),(6,annee_debut+1),(7,annee_debut+1),(8,annee_debut+1)]
    jf_nom = {0:'Lundi',1:'Mardi',2:'Mercredi',3:'Jeudi',4:'Vendredi'}

    wb_out = Workbook(); wb_out.remove(wb_out.active)
    total = 0

    for (mois, annee) in mois_scolaire:
        wb_tmp = load_workbook(template_path); ws_src = wb_tmp.active
        ws = wb_out.create_sheet(title=f"{MOIS_FR[mois]} {annee}")

        for row in ws_src.iter_rows():
            for cell in row:
                nc = ws.cell(row=cell.row, column=cell.column)
                nc.value = cell.value
                if cell.has_style:
                    nc.font = copy.copy(cell.font); nc.fill = copy.copy(cell.fill)
                    nc.border = copy.copy(cell.border); nc.alignment = copy.copy(cell.alignment)
                    nc.number_format = cell.number_format
        for cl, cd in ws_src.column_dimensions.items():
            ws.column_dimensions[cl].width = cd.width
        for ri, rd in ws_src.row_dimensions.items():
            ws.row_dimensions[ri].height = rd.height
        for mg in ws_src.merged_cells.ranges:
            ws.merge_cells(str(mg))

        for ri in range(1, 4):
            for ci2 in range(1, ws.max_column+1):
                v = ws.cell(row=ri, column=ci2).value
                if isinstance(v, str) and re.search(r'20\d\d', v):
                    ws.cell(row=ri, column=ci2).value = f"{MOIS_FR[mois].upper()} {annee}"

        nb_j = calendar.monthrange(annee, mois)[1]
        jours_ouvres = [datetime.date(annee, mois, j) for j in range(1, nb_j+1)
                        if datetime.date(annee, mois, j).weekday() < 5
                        and datetime.date(annee, mois, j) not in FERIES]

        row_idx = fdr
        for d in jours_ouvres:
            ds = d.strftime('%Y-%m-%d')
            ws.cell(row=row_idx, column=dlc).value = jf_nom[d.weekday()]
            ws.cell(row=row_idx, column=dlc+1).value = d.day
            for classe_nom, col_idx in cc.items():
                a_cours = any(
                    ds in pc['jours']
                    for pc in planning_classes
                    if pc['nom'] and (pc['nom'].strip().lower() in classe_nom.lower()
                                      or classe_nom.lower() in pc['nom'].strip().lower())
                )
                cell = ws.cell(row=row_idx, column=col_idx)
                if a_cours:
                    color = colors.get(classe_nom, 'FFD9D9D9')
                    if len(color) == 6: color = 'FF' + color
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                    total += 1
                else:
                    cell.fill = PatternFill(fill_type=None); cell.value = None
            row_idx += 1

    wb_out.save(output_path)
    return total, len(mois_scolaire)

# ─── Génération template vierge par mois ─────────────────────────────────────
def generer_template_mois(template_path, output_path, annee, mois):
    """
    Génère un template d'affichage vierge pour un mois donné.

    Algorithme :
    1. Calculer tous les jours Lun-Ven du mois (fériés inclus, les profs gèrent)
    2. Trouver slot_debut = weekday du 1er du mois (ou du 1er lundi si 1er=sam/dim)
       → supprimer les (slot_debut * 6) premières lignes de données (début de semaine vide)
    3. Écrire les labels de jours dans les slots restants
    4. Après le dernier jour réel, supprimer tous les slots vides restants

    Le template de référence a 25 slots → suffisant pour tous les mois.
    1 slot = 6 lignes. Séparateurs de semaine (h≈6) entre les vendredis et lundis.
    """


    # ── 1. Jours ouvrés du mois (Lun-Ven, fériés inclus) ─────────────────────
    jours_ouvres = []
    nb_jours = calendar.monthrange(annee, mois)[1]
    for j in range(1, nb_jours + 1):
        d = datetime.date(annee, mois, j)
        if d.weekday() < 5:
            jours_ouvres.append((JOURS_FR_LIST[d.weekday()], j))

    if not jours_ouvres:
        raise ValueError(f"Aucun jour ouvré pour {mois}/{annee}")

    # ── 2. Calculer les slots à supprimer au début ────────────────────────────
    # Si le 1er est un sam/dim, le mois commence dès le lundi suivant → slot_debut=0
    premier = datetime.date(annee, mois, 1)
    slot_debut = premier.weekday() if premier.weekday() < 5 else 0
    # Ex : 1er = Jeudi (3) → supprimer Lun/Mar/Mer = 3 slots = 18 lignes
    lignes_debut = slot_debut * 6

    # ── 3. Copier et ouvrir le template ──────────────────────────────────────
    shutil.copy(template_path, output_path)
    wb = load_workbook(output_path)
    ws = wb.active

    # Mettre à jour le titre
    nom_mois = MOIS_FR_UPPER[mois]
    for r in range(1, 4):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and 'SEPTEMBRE' in v.upper():
                ws.cell(row=r, column=c).value = v.replace(
                    "SEPTEMBRE 2026", f"{nom_mois} {annee}"
                )

    # ── 4. Supprimer les slots vides au début ─────────────────────────────────
    if lignes_debut > 0:
        DELETE_START = 7
        DELETE_COUNT = lignes_debut
        FIRST_KEPT   = DELETE_START + DELETE_COUNT

        # Sauvegarder les hauteurs AVANT suppression (bug openpyxl)
        saved_heights = {
            r: ws.row_dimensions[r].height
            for r in range(FIRST_KEPT, ws.max_row + 1)
        }
        ws.delete_rows(DELETE_START, DELETE_COUNT)

        # Ré-appliquer les hauteurs décalées
        for old_r, h in saved_heights.items():
            new_r = old_r - DELETE_COUNT
            if new_r >= DELETE_START and h is not None:
                ws.row_dimensions[new_r].height = h

        # Fix fusions optimisé (8x plus rapide : manipulation directe du set)
        from openpyxl.worksheet.cell_range import CellRange
        target_rows = {r for r in range(DELETE_START, ws.max_row + 1)
                       if (ws.row_dimensions[r].height is None or ws.row_dimensions[r].height >= 10)}
        for mg in list(ws.merged_cells.ranges):
            if mg.min_row == mg.max_row and mg.min_row in target_rows:
                try: ws.merged_cells.ranges.discard(mg)
                except Exception: pass
        for r in target_rows:
            for c1, c2 in ALL_MERGE_PAIRS:
                ws.merged_cells.ranges.add(
                    CellRange(f"{get_column_letter(c1)}{r}:{get_column_letter(c2)}{r}"))

    # ── 5. Détecter les slots disponibles après suppression ───────────────────
    jours_pos = []
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=2).value
        if isinstance(v, str) and any(j in v for j in JOURS_FR_LIST):
            jours_pos.append((r, r + 1))  # (label_row, num_row)

    # ── 6. Écrire les labels de jours dans les slots ──────────────────────────
    for i, (rl, rn) in enumerate(jours_pos):
        if i < len(jours_ouvres):
            label, num = jours_ouvres[i]
            for col in JOURS_COLS:
                ws.cell(row=rl, column=col).value = label
                ws.cell(row=rn, column=col).value = num
        else:
            # Slot excédentaire : effacer le label (sera supprimé ensuite)
            for col in JOURS_COLS:
                ws.cell(row=rl, column=col).value = None
                ws.cell(row=rn, column=col).value = None

    # ── 7. Supprimer les slots vides à la FIN ─────────────────────────────────
    # Chercher la ligne grise de fermeture (bg=FFC0C0C0) pour la préserver
    grey_closing_row = None
    for r in range(ws.max_row, 0, -1):
        fill = ws.cell(row=r, column=2).fill
        bg = fill.fgColor.rgb if fill and fill.fgColor and fill.fgColor.type == 'rgb' else None
        if bg == 'FFC0C0C0':
            grey_closing_row = r
            break

    if len(jours_pos) > len(jours_ouvres):
        last_label_row = jours_pos[len(jours_ouvres) - 1][0]
        last_bloc_end  = last_label_row + 4
        delete_from    = last_bloc_end + 1
        # S'arrêter juste avant la ligne grise de fermeture
        delete_until   = (grey_closing_row - 1) if grey_closing_row else ws.max_row
        if delete_from <= delete_until:
            ws.delete_rows(delete_from, delete_until - delete_from + 1)

    wb.save(output_path)
    return len(jours_ouvres)


# ─── Excel multi-feuilles (une feuille par mois) ─────────────────────────────
def _copier_feuille(ws_src, ws_dst):
    """
    Copie complète d'une feuille vers une autre (workbooks différents).
    Utilise copy.copy() pour les styles → pas de problème d'indices croisés.
    """
    for row in ws_src.iter_rows():
        for cell in row:
            nc = ws_dst.cell(row=cell.row, column=cell.column)
            nc.value = cell.value
            if cell.has_style:
                nc.font          = copy.copy(cell.font)
                nc.fill          = copy.copy(cell.fill)
                nc.border        = copy.copy(cell.border)
                nc.alignment     = copy.copy(cell.alignment)
                nc.number_format = cell.number_format
    for col, cd in ws_src.column_dimensions.items():
        ws_dst.column_dimensions[col].width = cd.width
    for r, rd in ws_src.row_dimensions.items():
        if rd.height:
            ws_dst.row_dimensions[r].height = rd.height
    seen = set()
    for mg in ws_src.merged_cells.ranges:
        k = str(mg)
        if k not in seen:
            seen.add(k)
            try:
                ws_dst.merge_cells(k)
            except Exception:
                pass


def generer_excel_multifeuilles(template_path, mois_liste, output_path):
    """
    Génère un Excel multi-feuilles (1 onglet par mois).
    Stratégie en 2 étapes :
      1. Générer chaque mois individuellement via generer_template_mois()
         (fonction existante, testée, fiable)
      2. Fusionner les fichiers dans un seul workbook via _copier_feuille()
         (copy.copy des styles → aucun problème d'indices croisés)
    """
    import tempfile

    wb_out = Workbook()
    wb_out.remove(wb_out.active)
    total_jours = 0

    for (annee, mois) in mois_liste:
        # Étape 1 : générer le mois dans un fichier temporaire
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name
        nb_jours = generer_template_mois(template_path, tmp_path, annee, mois)
        total_jours += nb_jours

        # Étape 2 : copier la feuille dans wb_out
        wb_tmp = load_workbook(tmp_path)
        ws_src = wb_tmp.active
        sheet_name = f"{MOIS_FR[mois][:4]} {annee}"
        ws_dst = wb_out.create_sheet(title=sheet_name)
        _copier_feuille(ws_src, ws_dst)
        wb_tmp.close()
        os.unlink(tmp_path)

    wb_out.save(output_path)
    return total_jours

# ─── Routes ───────────────────────────────────────────────────────────────────
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/generer', methods=['POST'])
def generer():
    sid = str(uuid.uuid4())[:8]; wd = os.path.join(UPLOAD_FOLDER, sid); os.makedirs(wd, exist_ok=True)
    def save(f, sub=''):
        d = os.path.join(wd, sub) if sub else wd; os.makedirs(d, exist_ok=True)
        p = os.path.join(d, f.filename); f.save(p); return p
    try:
        cf = request.files.getlist('classes'); df = request.files.getlist('dispos')
        ff = request.files.get('formateurs'); tf = request.files.get('template')
        mois = set(json.loads(request.form.get('mois', '[]')))
        if not all([cf, df, ff, tf, mois]): return jsonify({'error': 'Fichiers manquants'}), 400
        cp = [save(f, 'classes') for f in cf if f.filename]
        dp = [save(f, 'dispos')  for f in df if f.filename]
        fp = save(ff); tp = save(tf)
        pcs = [c for c in [parse_planning_classe(p) for p in cp] if c]
        dispos = [parse_disponibilite(p) for p in dp]
        aff = parse_tableau_formateurs(fp)
        assignment, stats, heures = assigner(pcs, dispos, aff)
        on = f"Planning_{'_'.join(m.replace('/','') for m in sorted(mois))}.xlsx"
        op = os.path.join(wd, on)
        ecrire_planning(tp, assignment, mois, op)
        return jsonify({
            'sessions_assignees': stats['assigned'], 'creneaux_sans_prof': stats['warn'],
            'formateurs_actifs': len(heures), 'classes': len(pcs),
            'mois': ', '.join(f"{MOIS_FR[int(m.split('/')[0])]} {m.split('/')[1]}" for m in mois),
            'fichier': on, 'session_id': sid,
            'heures_formateurs': {p: dict(m) for p, m in heures.items()}
        })
    except Exception as e:
        import traceback; return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500

@app.route('/generer-template', methods=['POST'])
def generer_template_colorie_route():
    sid = str(uuid.uuid4())[:8]; wd = os.path.join(UPLOAD_FOLDER, sid); os.makedirs(wd, exist_ok=True)
    def save(f, sub=''):
        d = os.path.join(wd, sub) if sub else wd; os.makedirs(d, exist_ok=True)
        p = os.path.join(d, f.filename); f.save(p); return p
    try:
        cf = request.files.getlist('classes'); tf = request.files.get('template')
        annee = int(request.form.get('annee', 2026))
        if not cf or not tf: return jsonify({'error': 'Fichiers manquants'}), 400
        cp = [save(f, 'classes') for f in cf if f.filename]; tp = save(tf)
        pcs = [c for c in [parse_planning_classe(p) for p in cp] if c]
        on = f"Template_Affichage_{annee}_{annee+1}.xlsx"; op = os.path.join(wd, on)
        total, nb = generer_template_colorie(tp, pcs, annee, op)
        return jsonify({
            'fichier': on, 'session_id': sid, 'nb_mois': nb,
            'cases_colories': total, 'classes': len(pcs),
            'noms_classes': [c['nom'] for c in pcs if c.get('nom')]
        })
    except Exception as e:
        import traceback; return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500


def _appliquer_mois_sur_feuille(ws, annee, mois, grey_row_override=None):
    """
    Applique les transformations d'un mois sur une feuille déjà copiée du template.
    Réutilise la même logique que generer_template_mois mais sur un ws existant.
    """
    # Titre
    nom_mois = MOIS_FR_UPPER[mois]
    for r in range(1, 4):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and 'SEPTEMBRE' in v.upper():
                ws.cell(row=r, column=c).value = v.replace(
                    "SEPTEMBRE 2026", f"{nom_mois} {annee}")

    # Jours ouvrés
    jours_ouvres = []
    for j in range(1, calendar.monthrange(annee, mois)[1] + 1):
        d = datetime.date(annee, mois, j)
        if d.weekday() < 5 and d not in FERIES:
            jours_ouvres.append((JOURS_FR_LIST[d.weekday()], j))

    # Suppression lignes début
    premier = datetime.date(annee, mois, 1)
    while premier.weekday() >= 5 or premier in FERIES:
        premier += datetime.timedelta(days=1)
    lignes_a_supprimer = premier.weekday() * 6

    if lignes_a_supprimer > 0:
        DELETE_START = 7
        DELETE_COUNT = lignes_a_supprimer
        FIRST_KEPT   = DELETE_START + DELETE_COUNT
        saved_heights = {r: ws.row_dimensions[r].height
                         for r in range(FIRST_KEPT, ws.max_row + 1)}
        ws.delete_rows(DELETE_START, DELETE_COUNT)
        for old_r, h in saved_heights.items():
            new_r = old_r - DELETE_COUNT
            if new_r >= DELETE_START and h is not None:
                ws.row_dimensions[new_r].height = h
        # Fix fusions optimisé
        from openpyxl.worksheet.cell_range import CellRange
        target_rows = {r for r in range(DELETE_START, ws.max_row + 1)
                       if (ws.row_dimensions[r].height is None or ws.row_dimensions[r].height >= 10)}
        for mg in list(ws.merged_cells.ranges):
            if mg.min_row == mg.max_row and mg.min_row in target_rows:
                try: ws.merged_cells.ranges.discard(mg)
                except Exception: pass
        for r in target_rows:
            for c1, c2 in ALL_MERGE_PAIRS:
                ws.merged_cells.ranges.add(
                    CellRange(f"{get_column_letter(c1)}{r}:{get_column_letter(c2)}{r}"))

    # Labels jours
    jours_pos = []
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=2).value
        if isinstance(v, str) and any(j in v for j in JOURS_FR_LIST):
            jours_pos.append((r, r + 1))

    for i, (rl, rn) in enumerate(jours_pos):
        if i < len(jours_ouvres):
            label, num = jours_ouvres[i]
            for col in JOURS_COLS:
                ws.cell(row=rl, column=col).value = label
                ws.cell(row=rn, column=col).value = num
        else:
            for col in JOURS_COLS:
                ws.cell(row=rl, column=col).value = None
                ws.cell(row=rn, column=col).value = None

    # Supprimer slots vides en fin (en préservant la ligne grise)
    # grey_row_override : position dans le template original (avant suppression lignes)
    # On la recalcule en tenant compte des lignes supprimées au début
    if grey_row_override is not None:
        grey_row = grey_row_override - lignes_a_supprimer
        if grey_row <= 0:
            grey_row = None
    else:
        grey_row = None
        for r in range(ws.max_row, 0, -1):
            try:
                fill = ws.cell(row=r, column=2).fill
                bg = (fill.fgColor.rgb
                      if fill and fill.fgColor and fill.fgColor.type == 'rgb'
                      else None)
                if bg == 'FFC0C0C0':
                    grey_row = r
                    break
            except Exception:
                pass
    if len(jours_pos) > len(jours_ouvres):
        last_rl = jours_pos[len(jours_ouvres) - 1][0]
        delete_from  = last_rl + 5
        delete_until = (grey_row - 1) if grey_row else ws.max_row
        if delete_from <= delete_until:
            ws.delete_rows(delete_from, delete_until - delete_from + 1)

    return len(jours_ouvres)

@app.route('/generer-template-vierge', methods=['POST'])
def generer_template_vierge_route():
    """
    Génère un ou plusieurs templates vierges (jusqu'à 12 mois) dans un ZIP.
    Paramètres POST :
      - template    : fichier .xlsx de référence
      - annee_debut : année de début
      - mois_debut  : mois de début (1-12)
      - annee_fin   : année de fin
      - mois_fin    : mois de fin (1-12)
    """
    import zipfile

    try:
        sid = str(uuid.uuid4())[:8]
        wd  = os.path.join(UPLOAD_FOLDER, sid)
        os.makedirs(wd, exist_ok=True)

        def save(f):
            p = os.path.join(wd, secure_filename(f.filename) if f.filename else 'template.xlsx')
            f.save(p)
            return p
        tf = request.files.get('template')
        if not tf:
            return jsonify({'error': 'Template manquant'}), 400

        annee_debut = int(request.form.get('annee_debut', 2026))
        mois_debut  = int(request.form.get('mois_debut',  1))
        annee_fin   = int(request.form.get('annee_fin',   2026))
        mois_fin    = int(request.form.get('mois_fin',    12))

        tp = save(tf)

        # Construire la liste ordonnée de mois à générer
        mois_liste = []
        a, m = annee_debut, mois_debut
        while (a, m) <= (annee_fin, mois_fin):
            mois_liste.append((a, m))
            m += 1
            if m > 12:
                m = 1; a += 1
            if len(mois_liste) > 12:
                break  # sécurité max 12 mois

        if not mois_liste:
            return jsonify({'error': 'Plage de mois invalide'}), 400

        fichiers_generes = []
        total_jours = 0

        for (annee, mois) in mois_liste:
            nom_mois = MOIS_FR_UPPER[mois]
            on = f"{nom_mois}_{annee}.xlsx"
            op = os.path.join(wd, on)
            nb_jours = generer_template_mois(tp, op, annee, mois)
            fichiers_generes.append(on)
            total_jours += nb_jours

        a_deb, m_deb = mois_liste[0]
        a_fin, m_fin = mois_liste[-1]
        label = f"{MOIS_FR_UPPER[m_deb]} {a_deb}" if len(mois_liste)==1 else f"{MOIS_FR_UPPER[m_deb]} {a_deb} → {MOIS_FR_UPPER[m_fin]} {a_fin}"
        format_sortie = request.form.get('format', 'zip')  # 'zip' ou 'excel'

        if format_sortie == 'excel' and len(mois_liste) > 1:
            # Excel multi-feuilles
            excel_name = f"Templates_{MOIS_FR_UPPER[m_deb]}_{a_deb}_au_{MOIS_FR_UPPER[m_fin]}_{a_fin}.xlsx"
            excel_path = os.path.join(wd, excel_name)
            generer_excel_multifeuilles(tp, mois_liste, excel_path)
            return jsonify({
                'fichier':    excel_name,
                'session_id': sid,
                'mois':       label,
                'nb_jours':   total_jours,
                'nb_mois':    len(mois_liste),
                'format':     'excel',
            })
        elif len(mois_liste) == 1:
            # Un seul mois : retourner directement le .xlsx
            on = fichiers_generes[0]
            return jsonify({
                'fichier':    on,
                'session_id': sid,
                'mois':       label,
                'nb_jours':   total_jours,
                'nb_mois':    1,
                'format':     'xlsx',
            })
        else:
            # ZIP (plusieurs fichiers séparés)
            zip_name = f"Templates_{MOIS_FR_UPPER[m_deb]}_{a_deb}_au_{MOIS_FR_UPPER[m_fin]}_{a_fin}.zip"
            zip_path = os.path.join(wd, zip_name)
            with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
                for fn in fichiers_generes:
                    zf.write(os.path.join(wd, fn), fn)
            return jsonify({
                'fichier':    zip_name,
                'session_id': sid,
                'mois':       label,
                'nb_jours':   total_jours,
                'nb_mois':    len(mois_liste),
                'format':     'zip',
            })

    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500

@app.route('/telecharger/<session_id>/<filename>')
def telecharger(session_id, filename):
    path = os.path.join(UPLOAD_FOLDER, session_id, filename)
    if not os.path.exists(path): return "Fichier introuvable", 404
    @after_this_request
    def cleanup(r):
        try: shutil.rmtree(os.path.join(UPLOAD_FOLDER, session_id))
        except: pass
        return r
    return send_file(path, as_attachment=True, download_name=filename)

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
